# backend/routes/attainment.py
import io
import os
from typing import Any, Dict, List, Optional
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from backend.core.auth import require_auth
from backend.database.user_models import User
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field
from sqlalchemy.ext.asyncio import AsyncSession

from backend.core.exceptions import OBEException
from backend.core.logger import get_logger
from backend.database.connection import get_db
from backend.services.attainment_service import AttainmentService
from backend.services.nba_report_service import NBAReportService

logger = get_logger(__name__)
router = APIRouter(prefix="/attainment", tags=["CO Attainment"])


# ── Request Models ──────────────────────────────────────────────────────

class StudentMarks(BaseModel):
    student_id: str = Field(example="USN001")
    student_name: str = Field(example="Alice Kumar")
    marks: Dict[str, Any] = Field(
        description="CO → component → marks (nested), OR exam → mark (flat exam-wise)",
        example={
            "CO1": {"Quiz": 8.0, "Unit Test": 18.0},
        }
    )


class MarksUploadRequest(BaseModel):
    students: List[StudentMarks] = Field(min_length=1)


# ── XLSX Parser ─────────────────────────────────────────────────────────

def _parse_one_sheet(ws, component_name: str) -> List[dict]:
    """
    Parse a single worksheet from the marks template.

    The sheet layout is:
      Row 0-2: title/info rows (skipped)
      Row 3:   header row  — 'SR', 'PRN', 'Student Name', 'Q1\n(/3)', 'Q2\n(/3)', ..., 'Total'
      Row 4:   max-marks row (skipped — PRN col contains 'Max →')
      Row 5+:  student data rows

    Returns a list of dicts:
        { "student_id": str, "student_name": str,
          "marks": { <component_name>: {"Total": float, "Q1": float, ...} } }
    """
    import re as _re

    all_rows = list(ws.iter_rows(values_only=True))

    # ── Find header row (contains 'PRN') ──────────────────────────────────
    header_row_idx = None
    for i, row in enumerate(all_rows):
        row_str = [str(v).strip().lower() if v is not None else '' for v in row]
        if 'prn' in row_str:
            header_row_idx = i
            break

    if header_row_idx is None:
        return []

    header_row = all_rows[header_row_idx]

    # ── Identify question sub-columns (Q1, Q2, …) and the Total column ──
    # These are columns from index 3 onward that contain "Q<n>" or "Total"
    SKIP_HEADER = {'sr', 'prn', 'roll no', 'student name', 'name', 'section', 'sec',
                   'grade', 'scaled', 'ca total', 'grand total'}
    question_cols = {}   # col_index → label ("Q1", "Q2", …)
    total_col = None     # col_index of the aggregated "Total" column

    for j, val in enumerate(header_row):
        if val is None:
            continue
        raw = str(val).strip().replace('\n', ' ').strip()
        raw_clean = raw.lower()
        # Strip marks suffix like "(/3)", "/3"
        label = _re.sub(r'\s*\(?\s*/\s*\d+\s*\)?\s*', '', raw).strip()

        if raw_clean in SKIP_HEADER or not label:
            continue
        if label.lower() == 'total':
            total_col = j
            continue
        # Detect question columns: Q1, Q2, Q3 … (any prefix+digit pattern)
        if _re.match(r'^Q\d+', label, _re.IGNORECASE):
            question_cols[j] = label
        # Also capture any other named sub-questions (e.g. Part A, Part B) — generic fallback
        elif j >= 3:
            question_cols[j] = label

    # ── Parse student rows ────────────────────────────────────────────────
    students = {}  # student_id → dict

    for row in all_rows[header_row_idx + 1:]:
        if not row or len(row) < 3:
            continue
        prn = row[1]    # Column B
        name = row[2]   # Column C

        if not prn or not name:
            continue
        prn_str = str(prn).strip()
        # Skip "Max →", section headers, averages
        if isinstance(prn, str) and not prn_str.replace('.', '').isdigit():
            continue
        if isinstance(name, str) and any(
            x in name.lower() for x in ['section', 'max', 'average', 'class avg', 'max marks']
        ):
            continue
        try:
            prn_int = int(float(prn_str))
            if prn_int < 100000:
                continue
        except (ValueError, TypeError):
            continue

        student_id = str(int(float(prn_str)))
        student_name = str(name).strip().replace('\xa0', '').strip()

        # Collect per-question marks
        sub_marks = {}
        for col_idx, q_label in question_cols.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val is None or str(val).strip() in ('', '—', 'N/A'):
                continue
            try:
                sub_marks[q_label] = float(val)
            except (TypeError, ValueError):
                pass

        # Prefer the pre-computed "Total" column; fall back to summing sub-marks
        if total_col is not None and total_col < len(row) and row[total_col] is not None:
            try:
                total_val = float(row[total_col])
            except (TypeError, ValueError):
                total_val = sum(sub_marks.values())
        else:
            total_val = sum(sub_marks.values())

        # Store Total AND per-question marks with integer keys (1, 2, 3…)
        # matching q_no so the master attainment frontend can use them directly.
        component_marks = {"Total": round(total_val, 4)}
        for q_label, q_mark in sub_marks.items():
            # Extract the question number: "Q1" → 1, "Q3" → 3
            import re as _re2
            m = _re2.match(r'Q(\d+)', str(q_label), _re2.IGNORECASE)
            if m:
                component_marks[int(m.group(1))] = round(q_mark, 4)

        if student_id in students:
            # Same student seen again (shouldn't happen within one sheet, but be safe)
            students[student_id]["marks"][component_name] = component_marks
        else:
            students[student_id] = {
                "student_id": student_id,
                "student_name": student_name,
                "marks": {component_name: component_marks},
            }

    return list(students.values())


def parse_marks_xlsx(file_bytes: bytes) -> List[dict]:
    """
    Parse a marks xlsx file.

    Supports two layouts:

    A) **Multi-sheet template** (preferred — used by the downloadable template):
       Each sheet is named after the evaluation component it covers:
         "Quiz 1", "Unit Test 1", "Unit Test 2", "Assignment", "End Semester"
       Within each sheet:
         - Info rows at top (skipped)
         - Header row: SR | PRN | Student Name | Q1(/n) | Q2(/n) | … | Total
         - Max-marks row (skipped)
         - Student rows
       Result: marks = { "Quiz 1": {"Total": 8.5, "Q1": 2.5, "Q2": 2.0, …},
                          "Unit Test 1": {"Total": 14.5, …}, … }

    B) **Single-sheet / CO-wise format** (legacy):
       Active sheet, columns: PRN | Name | CO1 | CO2 | …
       Result: marks = { "CO1": {"Total": 8.0}, "CO2": {"Total": 7.5}, … }

    B2) **Exam-wise single sheet** (legacy flat):
       Active sheet, columns: PRN | Name | Quiz | Unit Test | …
       Result: marks = { "Quiz": 8.5, "Unit Test": 14.0, … }  (_format="exam_wise")
    """
    import openpyxl
    import re as _re

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    # ── Layout A: multi-sheet template ────────────────────────────────────
    # Detect: workbook has > 1 sheet OR first sheet name != "Sheet1" / generic name
    GENERIC_NAMES = {'sheet', 'sheet1', 'marks', 'data', 'student marks'}
    sheet_names = [s.strip() for s in wb.sheetnames]
    is_multi_sheet = (
        len(sheet_names) > 1
        or (len(sheet_names) == 1 and sheet_names[0].lower() not in GENERIC_NAMES)
    )

    if is_multi_sheet:
        # Merge all sheets into one student dict, keyed by PRN
        merged: dict[str, dict] = {}   # student_id → {student_id, student_name, marks:{…}}

        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            component_name = sheet_name.strip()
            sheet_students = _parse_one_sheet(ws, component_name)

            for s in sheet_students:
                sid = s["student_id"]
                if sid not in merged:
                    merged[sid] = {
                        "student_id": sid,
                        "student_name": s["student_name"],
                        "marks": {},
                    }
                merged[sid]["marks"].update(s["marks"])

        students = list(merged.values())
        if students:
            return students
        # Fall through to legacy single-sheet logic if nothing parsed

    # ── Layout B / B2: single-sheet (legacy) ─────────────────────────────
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))

    # Find header row
    header_row_idx = None
    co_columns = {}   # col_index → CO name  (e.g. "CO1")
    for i, row in enumerate(all_rows):
        row_str = [str(v).strip().lower() if v is not None else '' for v in row]
        if 'prn' in row_str:
            header_row_idx = i
            for j, val in enumerate(row):
                if val and str(val).strip().upper().startswith('CO'):
                    raw = str(val).strip().upper().split('\n')[0].strip()
                    m = _re.match(r'(CO\d+)', raw)
                    if m:
                        co_columns[j] = m.group(1)
            break

    if header_row_idx is None:
        raise ValueError("Could not find header row with 'PRN' column in the xlsx file.")

    # Exam-wise columns (non-CO, non-skip)
    header_row = all_rows[header_row_idx]
    exam_columns = {}
    SKIP_COLS = {'sr', 'sr no', 'sr. no', 'sr.no', 'prn', 'roll no', 'student name', 'name',
                 'sec', 'section', 'ca total', 'grand total', 'grade', 'scaled', 'total',
                 'max', 'max →', 'max->', 'max marks'}

    for j, val in enumerate(header_row):
        if val is None:
            continue
        raw = str(val).strip().replace('\n', ' ').strip()
        raw_lower = raw.lower()
        # Strip marks suffix
        label = _re.sub(r'\s*\(?\s*/\s*\d+\s*\)?\s*', '', raw).strip()
        label_lower = label.lower()

        if any(s in raw_lower for s in SKIP_COLS) or not label:
            continue
        if raw.upper().startswith('CO') and raw[2:3].isdigit():
            continue
        # Skip Q1/Q2/… style sub-question columns (these are sub-marks, not components)
        if _re.match(r'^Q\d+', label, _re.IGNORECASE):
            continue
        if label and not label.replace('.', '').replace('/', '').isdigit():
            if label_lower not in SKIP_COLS:
                exam_columns[j] = label

    students = []
    for row in all_rows[header_row_idx + 1:]:
        if not row or len(row) < 3:
            continue
        prn = row[1]
        name = row[2]
        if not prn or not name:
            continue
        prn_str = str(prn).strip()
        if isinstance(prn, str) and not prn_str.replace('.', '').isdigit():
            continue
        if isinstance(name, str) and any(
            x in name.lower() for x in ['section', 'max', 'average', 'class avg', 'max marks']
        ):
            continue
        try:
            prn_int = int(float(prn_str))
            if prn_int < 100000:
                continue
        except (ValueError, TypeError):
            continue

        student_id = str(int(float(str(prn)))).strip()
        student_name = str(name).strip().replace('\xa0', '').strip()

        marks = {}
        if exam_columns:
            for col_idx, exam_name in exam_columns.items():
                val = row[col_idx] if col_idx < len(row) else None
                if val is None or str(val).strip() in ('', '—', 'N/A'):
                    continue
                try:
                    marks[exam_name] = float(val)
                except (TypeError, ValueError):
                    pass
            if marks:
                students.append({
                    "student_id": student_id,
                    "student_name": student_name,
                    "marks": marks,
                    "_format": "exam_wise",
                })
                continue
        if co_columns:
            for col_idx, co_name in co_columns.items():
                val = row[col_idx] if col_idx < len(row) else None
                try:
                    marks[co_name] = {"Total": float(val) if val is not None else 0.0}
                except (TypeError, ValueError):
                    marks[co_name] = {"Total": 0.0}
        else:
            co_idx = 1
            for j in range(3, len(row)):
                val = row[j]
                if val is not None:
                    try:
                        marks[f"CO{co_idx}"] = {"Total": float(val)}
                        co_idx += 1
                    except (TypeError, ValueError):
                        pass

        if student_id and student_name:
            students.append({
                "student_id": student_id,
                "student_name": student_name,
                "marks": marks if marks else {"CO1": {"Total": 0.0}},
            })

    return students


# ── Routes ─────────────────────────────────────────────────────────────

@router.post("/marks/{course_id}/xlsx", status_code=201)
async def upload_marks_xlsx(
    course_id: int,
    file: UploadFile = File(..., description="Student marks Excel file (.xlsx)"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """
    Upload student marks from an Excel (.xlsx) file.
    Expected columns: PRN (col B), Student Name (col C), CO marks (col D+).
    Handles institute header rows, section separators, and formula cells automatically.
    Replaces all previous marks for this course.
    """
    logger.info(f"XLSX marks upload for course_id={course_id}, file={file.filename}")
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported.")
    try:
        file_bytes = await file.read()
        students = parse_marks_xlsx(file_bytes)
        if not students:
            raise HTTPException(status_code=400, detail="No student records found in the file. Check the format.")
        logger.info(f"Parsed {len(students)} students from xlsx")
        # Detect if this is an exam-wise file (flat marks dict)
        is_exam_wise = any(s.get("_format") == "exam_wise" for s in students)
        # Clean up the _format key and normalize marks for saving
        cleaned_students = []
        for s in students:
            fmt = s.pop("_format", None)
            if fmt == "exam_wise":
                # Store flat exam marks directly — frontend handles display
                cleaned_students.append({
                    "student_id": s["student_id"],
                    "student_name": s["student_name"],
                    "marks": s["marks"],  # {"Quiz": 8.5, "Unit Test": 7.0, ...}
                })
            else:
                cleaned_students.append(s)
        service = AttainmentService(db)
        result = await service.save_marks(course_id, cleaned_students)
        return {
            "status": "success",
            "data": result,
            "parsed_students": len(cleaned_students),
            "format": "exam_wise" if is_exam_wise else "co_wise",
        }
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except OBEException as e:
        logger.error(f"Marks upload error: {e.message}")
        raise HTTPException(status_code=e.status_code, detail=e.message)
    except Exception:
        logger.exception(f"Unexpected error uploading xlsx marks for course {course_id}")
        raise HTTPException(status_code=500, detail="Internal server error")


@router.post("/marks/{course_id}", status_code=201)
async def upload_marks(
    course_id: int,
    request: MarksUploadRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """
    Upload student marks per CO per evaluation component (JSON body).
    Replaces all previous marks for this course.
    """
    logger.info(f"Marks upload for course_id={course_id}, students={len(request.students)}")
    try:
        service = AttainmentService(db)
        result = await service.save_marks(
            course_id,
            [s.dict() for s in request.students],
        )
        return {"status": "success", "data": result}
    except OBEException as e:
        logger.error(f"Marks upload error: {e.message}")
        raise HTTPException(status_code=e.status_code, detail=e.message)
    except Exception:
        logger.exception(f"Unexpected error uploading marks for course {course_id}")
        raise HTTPException(status_code=500, detail="Internal server error")


@router.get("/marks/{course_id}")
async def get_marks(
    course_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """
    Return all student marks records for a course as a list of dicts.
    Used by the frontend to display and edit uploaded marks.
    """
    from sqlalchemy import select as _select
    from backend.database.models import COAttainment
    logger.info(f"Fetching marks for course_id={course_id}")
    try:
        result = await db.execute(
            _select(COAttainment).where(COAttainment.course_id == course_id)
        )
        records = result.scalars().all()
        return {"status": "success", "data": [r.to_dict() for r in records], "total": len(records)}
    except Exception:
        logger.exception(f"Error fetching marks for course {course_id}")
        raise HTTPException(status_code=500, detail="Internal server error")


@router.put("/marks/{course_id}/student/{student_id}", status_code=200)
async def update_student_marks(
    course_id: int,
    student_id: str,
    request: MarksUploadRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """
    Update marks for a single student in a course.
    Finds the existing COAttainment record by course_id + student_id and patches marks.
    """
    from sqlalchemy import select as _select
    from backend.database.models import COAttainment
    logger.info(f"Updating marks for course_id={course_id} student_id={student_id}")
    try:
        result = await db.execute(
            _select(COAttainment).where(
                COAttainment.course_id == course_id,
                COAttainment.student_id == student_id,
            )
        )
        rec = result.scalar_one_or_none()
        if not rec:
            raise HTTPException(status_code=404, detail=f"No marks found for student_id={student_id}")
        student_data = request.students[0]
        rec.student_name = student_data.student_name
        rec.marks = student_data.marks
        await db.commit()
        await db.refresh(rec)
        return {"status": "success", "data": rec.to_dict()}
    except HTTPException:
        raise
    except Exception:
        logger.exception(f"Error updating marks for course {course_id} student {student_id}")
        raise HTTPException(status_code=500, detail="Internal server error")


@router.get("/documents/{course_id}")
async def list_documents(course_id: int, current_user: User = Depends(require_auth)):
    """
    List all generated documents for a course (session plan, evaluation plan,
    attainment report, NBA report, question papers).
    Returns filename, category, size, download URL.
    """
    from backend.core.storage import get_storage
    storage = get_storage()
    docs = []
    checks = [
        ("session_plans",     f"session_plan_{course_id}.docx",    "Session Plan",        f"/session-plan/download/{course_id}",            "📅"),
        ("evaluation_plans",  f"eval_plan_{course_id}.docx",       "Evaluation Plan",     f"/evaluation-plan/download/{course_id}",          "📋"),
        ("attainment_reports",f"attainment_report_{course_id}.docx","Attainment Report",  f"/attainment/download/{course_id}",               "📊"),
        ("nba_reports",       f"nba_report_{course_id}.pdf",        "NBA/NAAC PDF",        f"/attainment/nba-report/download/{course_id}",    "🏆"),
        ("question_papers",   f"qpaper_{course_id}_latest.docx",    "Question Paper (.docx)",f"/questions/paper/download/{course_id}/docx",  "❓"),
        ("question_papers",   f"qpaper_{course_id}_latest.pdf",     "Question Paper (.pdf)", f"/questions/paper/download/{course_id}/pdf",   "❓"),
    ]
    for category, filename, label, url, icon in checks:
        p = storage.get_path(category, filename)
        if p and p.exists():
            stat = p.stat()
            docs.append({
                "label": label,
                "filename": filename,
                "category": category,
                "icon": icon,
                "size_kb": round(stat.st_size / 1024, 1),
                "download_url": url,
                "generated_at": stat.st_mtime,
            })
        else:
            # Also try listing by prefix for question papers which may have timestamps
            if category == "question_papers":
                files = storage.list_files(category, prefix=f"qpaper_{course_id}_")
                for f in files:
                    stat = f.stat()
                    ext = f.suffix
                    lbl = f"Question Paper ({ext})"
                    dl_url = f"/questions/paper/download/{course_id}/docx" if ext == ".docx" else f"/questions/paper/download/{course_id}/pdf"
                    docs.append({
                        "label": lbl, "filename": f.name, "category": category,
                        "icon": "❓", "size_kb": round(stat.st_size / 1024, 1),
                        "download_url": dl_url, "generated_at": stat.st_mtime,
                    })
    return {"status": "success", "data": docs, "total": len(docs)}


@router.get("/calculate/{course_id}")
async def calculate_attainment(
    course_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """
    Calculate and return CO + PO attainment as JSON (no file generated).
    Use this to preview attainment before generating the report.
    """
    try:
        service = AttainmentService(db)
        result = await service.calculate(course_id)
        return {"status": "success", "data": result}
    except OBEException as e:
        raise HTTPException(status_code=e.status_code, detail=e.message)
    except Exception:
        logger.exception(f"Error calculating attainment for course {course_id}")
        raise HTTPException(status_code=500, detail="Internal server error")


@router.post("/report/{course_id}", status_code=201)
async def generate_attainment_report(
    course_id: int,
    template: Optional[UploadFile] = File(None, description="Optional .docx template to guide report structure"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """
    Generate a formatted Word document attainment report.
    Requires marks to be uploaded first via POST /attainment/marks/{course_id}.
    Optionally accepts a .docx template — if provided, the report follows that template's structure.
    """
    logger.info(f"Attainment report requested for course_id={course_id}, template={'yes' if template else 'no'}")
    try:
        service = AttainmentService(db)
        if template and template.filename:
            template_bytes = await template.read()
            result = await service.generate_report_from_template(course_id, template_bytes)
        else:
            result = await service.generate_report(course_id)
        return {"status": "success", "data": result}
    except OBEException as e:
        logger.error(f"Attainment report error: {e.message}")
        raise HTTPException(status_code=e.status_code, detail=e.message)
    except Exception:
        logger.exception(f"Unexpected error generating attainment report for course {course_id}")
        raise HTTPException(status_code=500, detail="Internal server error")


@router.get("/download/{course_id}")
async def download_attainment_report(course_id: int, current_user: User = Depends(require_auth)):
    """Download the generated attainment report Word document."""
    filepath = AttainmentService.get_filepath(course_id)
    if not os.path.exists(filepath):
        raise HTTPException(
            status_code=404,
            detail=f"Report not found for course_id={course_id}. Run POST /attainment/report/{course_id} first.",
        )
    return FileResponse(
        path=filepath,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        filename=f"attainment_report_{course_id}.docx",
    )


# ── Day 6: NBA/NAAC PDF Routes ──────────────────────────────────────────

@router.post("/nba-report/{course_id}", status_code=201)
async def generate_nba_report(
    course_id: int,
    template: Optional[UploadFile] = File(None, description="Optional .docx template to guide report structure"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """
    Generate a full NBA/NAAC-format PDF report with:
    - Section A: NBA CO attainment + gap analysis + bar charts
    - Section B: NAAC PO attainment + level descriptors
    - Section C: AI-generated recommendations
    - Section D: CO-PO correlation matrix

    Requires marks to be uploaded first via POST /attainment/marks/{course_id}.
    Optionally accepts a .docx template — if provided, the PDF follows that template's section structure.
    """
    logger.info(f"NBA/NAAC PDF report requested for course_id={course_id}, template={'yes' if template else 'no'}")
    try:
        service = NBAReportService(db)
        if template and template.filename:
            template_bytes = await template.read()
            result = await service.generate_pdf_from_template(course_id, template_bytes)
        else:
            result = await service.generate_pdf(course_id)
        return {"status": "success", "data": result}
    except OBEException as e:
        logger.error(f"NBA report error: {e.message}")
        raise HTTPException(status_code=e.status_code, detail=e.message)
    except Exception:
        logger.exception(f"Unexpected error generating NBA report for course {course_id}")
        raise HTTPException(status_code=500, detail="Internal server error")


@router.get("/nba-report/download/{course_id}")
async def download_nba_report(course_id: int, current_user: User = Depends(require_auth)):
    """Download the generated NBA/NAAC PDF report."""
    filepath = NBAReportService.get_filepath(course_id)
    if not os.path.exists(filepath):
        raise HTTPException(
            status_code=404,
            detail=f"NBA report not found for course_id={course_id}. Run POST /attainment/nba-report/{course_id} first.",
        )
    return FileResponse(
        path=filepath,
        media_type="application/pdf",
        filename=f"nba_report_{course_id}.pdf",
    )


@router.get("/nba-report/gap-analysis/{course_id}")
async def get_gap_analysis(
    course_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """
    Preview the CO-PO gap analysis as JSON without generating the PDF.
    Shows which COs and POs are at risk (below target threshold).
    """
    try:
        service = NBAReportService(db)
        result = await service.gap_analysis(course_id)
        return {"status": "success", "data": result}
    except OBEException as e:
        raise HTTPException(status_code=e.status_code, detail=e.message)
    except Exception:
        logger.exception(f"Error running gap analysis for course {course_id}")
        raise HTTPException(status_code=500, detail="Internal server error")