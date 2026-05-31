# backend/services/co_po_template_service.py
"""
Generates a pre-filled CO-PO Attainment Excel workbook (.xlsx) for a course.
Matches the SIT template structure exactly.

Sheet layout (fixed rows):
  Course_Info   — labels in col A, values in col C (rows 1-11)
  Roll_List     — header block; students from row 8
  CO_List       — header block; rubric rows 7-10, CO table from row 13
  ESE_QP / CA{n}_QP  — header block; col headers row 7; questions from row 8
  ESE_MKS / CA{n}_Marks — header block; col headers row 7; "Marks" row 8;
                            max row 9; students from row 10; summary rows after last student
  Final_CO_Attn — references CA_Marks summary rows and CO_List
  PO_Attainment — CO-PO matrix + attainment formulas

KEY DESIGN DECISION: All cells where we already know the value at generation
time are written as HARDCODED VALUES, not formulas. This ensures the file
opens correctly in any viewer (openpyxl data_only, Google Sheets, LibreOffice)
without needing Excel formula recalculation. Formulas are only used for
cross-sheet summary/attainment calculations that must stay dynamic.
"""

import os
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from backend.core.logger import get_logger
from backend.core.storage import get_storage
from backend.services.course_service import CourseService

logger = get_logger(__name__)

_CATEGORY = "co_po_templates"

# ── Colours ──────────────────────────────────────────────────────────────────
_NAVY_HEX    = "1F3864"
_LIGHT_HEX   = "D6DCE4"
_GREEN_HEX   = "E2EFDA"
_ORANGE_HEX  = "FCE4D6"
_YELLOW_HEX  = "FFFF00"
_SKYBLUE_HEX = "BDD7EE"
_TEAL_HEX    = "00B0F0"
_LIME_HEX    = "92D050"
_PINK_HEX    = "FF99CC"

_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
_SUBHDR_FONT = Font(name="Calibri", bold=True, size=10)
_BODY_FONT   = Font(name="Calibri", size=10)
_NAVY_FILL   = PatternFill("solid", fgColor=_NAVY_HEX)
_LIGHT_FILL  = PatternFill("solid", fgColor=_LIGHT_HEX)
_GREEN_FILL  = PatternFill("solid", fgColor=_GREEN_HEX)
_ORANGE_FILL = PatternFill("solid", fgColor=_ORANGE_HEX)
_YELLOW_FILL = PatternFill("solid", fgColor=_YELLOW_HEX)
_SKYBLUE_FILL= PatternFill("solid", fgColor=_SKYBLUE_HEX)
_TEAL_FILL   = PatternFill("solid", fgColor=_TEAL_HEX)
_LIME_FILL   = PatternFill("solid", fgColor=_LIME_HEX)
_PINK_FILL   = PatternFill("solid", fgColor=_PINK_HEX)

_THICK  = Side(style="medium")
_THIN   = Side(style="thin")
_BORDER = Border(left=_THICK, right=_THICK, top=_THICK, bottom=_THICK)
_THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# Maximum question columns when no QP is provided
_MAX_Q = 30


def _c(ws, row, col, value=None, font=None, fill=None, align=None, border=True, bold=False, number_format=None):
    """Write a styled cell with bold black border by default."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font or Font(name="Calibri", size=10, bold=bold)
    if fill:
        cell.fill = fill
    cell.border = _BORDER if border else _THIN_BORDER
    cell.alignment = align or _CENTER
    if number_format:
        cell.number_format = number_format
    return cell


def _navy(ws, row, col, value, align=None):
    return _c(ws, row, col, value, font=_HEADER_FONT, fill=_NAVY_FILL, align=align or _CENTER)


def _merge(ws, r1, r2, c1, c2, value=None, font=None, fill=None, align=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=value)
    cell.font = font or _SUBHDR_FONT
    if fill:
        cell.fill = fill
    cell.border = _BORDER
    cell.alignment = align or _CENTER
    return cell


def _qp_sheet_name(name):
    """Return sheet name quoted with single quotes if it contains spaces."""
    if " " in name:
        return f"'{name}'"
    return name


# ─────────────────────────────────────────────────────────────────────────────
# Standard 6-row course header block (used in every sheet)
# ─────────────────────────────────────────────────────────────────────────────
def _course_header_block(ws, course):
    dept_text = f"Department of : {course.department}"
    ws.merge_cells("A1:J1")
    c = ws.cell(row=1, column=1, value=dept_text)
    c.font = Font(name="Calibri", bold=True, size=12)
    c.fill = _SKYBLUE_FILL
    c.border = _BORDER
    c.alignment = _CENTER

    ws.merge_cells("A2:J2")
    c = ws.cell(row=2, column=1, value="CO Attainment")
    c.font = Font(name="Calibri", bold=True, size=11)
    c.fill = _TEAL_FILL
    c.border = _BORDER
    c.alignment = _CENTER

    academic_year = course.academic_year
    batch = getattr(course, "batch", course.academic_year)
    exam_season = getattr(course, "exam_season", "")

    _c(ws, 4, 1, "Academic Year",      bold=True, fill=_LIGHT_FILL, align=_LEFT)
    _c(ws, 4, 3, academic_year,         fill=_YELLOW_FILL, align=_CENTER)
    _c(ws, 4, 5, "Batch",              bold=True, fill=_LIGHT_FILL, align=_LEFT)
    _c(ws, 4, 6, batch,                fill=_YELLOW_FILL, align=_CENTER)
    _c(ws, 4, 8, "Examination Season", bold=True, fill=_LIGHT_FILL, align=_LEFT)
    _c(ws, 4, 10, exam_season,          fill=_YELLOW_FILL, align=_CENTER)

    _c(ws, 5, 1, "Course Name",        bold=True, fill=_LIGHT_FILL, align=_LEFT)
    _c(ws, 5, 3, course.course_name,   fill=_YELLOW_FILL, align=_LEFT)
    _c(ws, 5, 8, "Course Code",        bold=True, fill=_LIGHT_FILL, align=_LEFT)
    _c(ws, 5, 10, course.course_code,  fill=_YELLOW_FILL, align=_CENTER)

    return 7  # first usable content row


# ─────────────────────────────────────────────────────────────────────────────
# Course_Info sheet
# ─────────────────────────────────────────────────────────────────────────────
def _build_course_info(wb, course):
    ws = wb.create_sheet("Course_Info")

    ws.merge_cells("A1:K1")
    c = ws.cell(row=1, column=1, value=f"Department of : {course.department}")
    c.font = Font(name="Calibri", bold=True, size=12)
    c.fill = _SKYBLUE_FILL
    c.border = _BORDER
    c.alignment = _CENTER

    ws.merge_cells("A2:K2")
    c = ws.cell(row=2, column=1, value="CO Attainment")
    c.font = Font(name="Calibri", bold=True, size=11)
    c.fill = _TEAL_FILL
    c.border = _BORDER
    c.alignment = _CENTER

    labels = ["Academic Year", "Batch", "Examination Season", "Course Name",
              "Course Code", "Semester", "Credit", "Faculty Name"]
    vals = [
        course.academic_year,
        getattr(course, "batch", course.academic_year),
        getattr(course, "exam_season", ""),
        course.course_name,
        course.course_code,
        course.semester,
        course.credits,
        course.faculty_name,
    ]
    fills = [_LIGHT_FILL, _LIGHT_FILL, _LIGHT_FILL, _GREEN_FILL,
             _GREEN_FILL, _LIGHT_FILL, _LIGHT_FILL, _ORANGE_FILL]
    for i, (lbl, val, fill) in enumerate(zip(labels, vals, fills)):
        r = i + 4
        _c(ws, r, 1, lbl, bold=True, fill=_LIGHT_FILL, align=_LEFT)
        _c(ws, r, 3, val, fill=fill, align=_LEFT)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["C"].width = 50


# ─────────────────────────────────────────────────────────────────────────────
# Roll_List
# ─────────────────────────────────────────────────────────────────────────────
def _build_roll_list(wb, course, students):
    ws = wb.create_sheet("Roll_List")
    r = _course_header_block(ws, course)

    headers = ["Sr. No.", "Seat No", "PRN", "Name of the Student", "Section"]
    for ci, h in enumerate(headers, 1):
        _c(ws, r, ci, h, font=_HEADER_FONT, fill=_NAVY_FILL)
    r += 1

    row_fills = [_GREEN_FILL, _LIGHT_FILL]
    for idx, s in enumerate(students, 1):
        fill = row_fills[idx % 2]
        prn_val = str(s["prn"]) if s["prn"] is not None else ""
        _c(ws, r, 1, idx,                    fill=fill, align=_CENTER)
        _c(ws, r, 2, "",                      fill=fill, align=_CENTER)
        _c(ws, r, 3, prn_val,                fill=fill, align=_CENTER)
        _c(ws, r, 4, s["name"],              fill=fill, align=_LEFT)
        _c(ws, r, 5, s.get("section", ""),   fill=fill, align=_CENTER)
        r += 1

    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 32
    ws.column_dimensions["A"].width = 8


# ─────────────────────────────────────────────────────────────────────────────
# CO_List
# ─────────────────────────────────────────────────────────────────────────────
def _build_co_list(wb, course):
    ws = wb.create_sheet("CO_List")
    _course_header_block(ws, course)

    r = 7
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    c = ws.cell(row=r, column=1, value="Rubric for deciding level of attainment")
    c.font = Font(name="Calibri", bold=True, size=10)
    c.fill = _ORANGE_FILL
    c.border = _BORDER
    c.alignment = _CENTER
    _c(ws, r, 9,  "Range", bold=True, fill=_NAVY_FILL, font=_HEADER_FONT)
    _c(ws, r, 10, "",      fill=_NAVY_FILL)
    _c(ws, r, 11, "Level", bold=True, fill=_NAVY_FILL, font=_HEADER_FONT)
    r += 1

    rubric_data = [
        ("If the percentage of students is less than equal to 40% secured >= 60%  marks ", "<= 40%",        1, _LIGHT_FILL),
        ("If the percentage of students is > 40% and  < 70% secured >= 60% marks ",        "> 40% & < 70%", 2, _GREEN_FILL),
        ("If the percentage of students is greater than or equal to  70% secured >= 60%  marks ", ">=70%",  3, _SKYBLUE_FILL),
    ]
    for text_val, rng, lvl, rfill in rubric_data:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        c = ws.cell(row=r, column=1, value=text_val)
        c.font = _BODY_FONT
        c.fill = rfill
        c.border = _BORDER
        c.alignment = _LEFT
        _c(ws, r, 9,  rng, fill=rfill,        align=_CENTER)
        _c(ws, r, 10, "",  fill=rfill)
        _c(ws, r, 11, lvl, fill=_YELLOW_FILL, align=_CENTER, bold=True)
        r += 1

    r = 13
    _c(ws, r, 1,  "CO No",    bold=True, fill=_NAVY_FILL, font=_HEADER_FONT, align=_CENTER)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
    c = ws.cell(row=r, column=2, value="Statement")
    c.font = _HEADER_FONT
    c.fill = _NAVY_FILL
    c.border = _BORDER
    c.alignment = _LEFT
    _c(ws, r, 10, "Target (% of maximum marks)", bold=True, fill=_NAVY_FILL, font=_HEADER_FONT, align=_CENTER)
    r += 1

    co_fills = [_GREEN_FILL, _LIGHT_FILL, _YELLOW_FILL, _ORANGE_FILL, _PINK_FILL]
    for i, co in enumerate(course.cos):
        fill = co_fills[i % len(co_fills)]
        _c(ws, r, 1, co["co_id"], fill=fill, bold=True, align=_CENTER)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
        c = ws.cell(row=r, column=2, value=co["statement"])
        c.font = _BODY_FONT
        c.fill = fill
        c.border = _BORDER
        c.alignment = _LEFT
        _c(ws, r, 10, 60, fill=_YELLOW_FILL, bold=True, align=_CENTER)
        r += 1

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 70


# ─────────────────────────────────────────────────────────────────────────────
# QP sheet (ESE_QP or CA{n}_QP)
# ─────────────────────────────────────────────────────────────────────────────
def _build_qp_sheet(wb, sheet_name, course, ca_label, questions=None):
    ws = wb.create_sheet(sheet_name)
    _course_header_block(ws, course)

    r = 7
    hdr_labels = ["Q. No", "Question", "", "", "", "", "Marks",
                  "CO Map to question", "BL", "",
                  "CO", "Marks for the CO", "Percentage", "",
                  "Bloom's Taxonomy Level (BL)", "", "Marks for BL", "Percentage"]
    hdr_fills  = [_NAVY_FILL, _NAVY_FILL, _NAVY_FILL, _NAVY_FILL, _NAVY_FILL,
                  _NAVY_FILL, _TEAL_FILL, _ORANGE_FILL, _LIME_FILL, _NAVY_FILL,
                  _NAVY_FILL, _SKYBLUE_FILL, _SKYBLUE_FILL, _NAVY_FILL,
                  _NAVY_FILL, _NAVY_FILL, _SKYBLUE_FILL, _SKYBLUE_FILL]
    for ci, (h, hf) in enumerate(zip(hdr_labels, hdr_fills), 1):
        _c(ws, r, ci, h, font=_HEADER_FONT, fill=hf)
    ws.row_dimensions[r].height = 30

    bloom = [("L1", "Remembering"), ("L2", "Understanding"), ("L3", "Applying"),
             ("L4", "Analyzing"),   ("L5", "Evaluating"),    ("L6", "Creating")]
    bl_row_fills = [_LIGHT_FILL, _GREEN_FILL, _YELLOW_FILL,
                    _ORANGE_FILL, _PINK_FILL, _LIME_FILL]

    # Pre-compute CO summary and BL summary from questions (Python SUMIF)
    cos = [c["co_id"] for c in course.cos]
    co_marks_sum = {}   # co_id -> total marks assigned to it
    bl_marks_sum = {}   # bl_label -> total marks assigned to it
    total_marks_qp = 0
    if questions:
        for q in questions:
            marks = q.get("marks") or 0
            try:
                marks = float(marks)
            except (TypeError, ValueError):
                marks = 0
            co_id = q.get("co_id", "")
            if co_id:
                co_marks_sum[co_id] = co_marks_sum.get(co_id, 0.0) + marks
            bl_raw = str(q.get("bloom_level", ""))
            bl_str = bl_raw if bl_raw.startswith("L") else (f"L{bl_raw}" if bl_raw else "")
            if bl_str:
                bl_marks_sum[bl_str] = bl_marks_sum.get(bl_str, 0.0) + marks
            total_marks_qp += marks

    for i, (bl_label, bl_name) in enumerate(bloom):
        sr = r + 1 + i   # rows 8-13
        bf = bl_row_fills[i]

        # CO summary: col K = CO name from course COs list, L = marks for that CO, M = %
        co_id_for_row = cos[i] if i < len(cos) else ""
        co_marks = co_marks_sum.get(co_id_for_row, "") if co_id_for_row else ""
        co_pct   = ""
        if co_marks and total_marks_qp > 0:
            try:
                co_pct = round(float(co_marks) / total_marks_qp * 100, 2)
            except (TypeError, ValueError):
                co_pct = ""

        _c(ws, sr, 11, co_id_for_row if co_id_for_row else "", fill=bf, align=_CENTER)
        _c(ws, sr, 12, co_marks if co_marks else "", fill=bf, align=_CENTER)
        _c(ws, sr, 13, co_pct if co_pct != "" else "", fill=bf, align=_CENTER, number_format="0.00")

        # BL summary: col O = BL code, P = name, Q = marks for BL, R = %
        bl_marks = bl_marks_sum.get(bl_label, "")
        bl_pct   = ""
        if bl_marks and total_marks_qp > 0:
            try:
                bl_pct = round(float(bl_marks) / total_marks_qp * 100, 2)
            except (TypeError, ValueError):
                bl_pct = ""

        _c(ws, sr, 15, bl_label, fill=bf, bold=True, align=_CENTER)
        _c(ws, sr, 16, bl_name,  fill=bf, align=_LEFT)
        _c(ws, sr, 17, bl_marks if bl_marks else "", fill=bf, align=_CENTER)
        _c(ws, sr, 18, bl_pct   if bl_pct   != "" else "", fill=bf, align=_CENTER, number_format="0.00")
        ws.row_dimensions[sr].height = 22

    q_row = r + 1
    q_row_fills = [_GREEN_FILL, _LIGHT_FILL]
    if questions:
        for qi, q in enumerate(questions):
            qf = q_row_fills[qi % 2]
            _c(ws, q_row, 1, q.get("q_no", qi + 1), fill=qf, bold=True, align=_CENTER)
            ws.merge_cells(start_row=q_row, start_column=2, end_row=q_row, end_column=6)
            c = ws.cell(row=q_row, column=2, value=q.get("question_text", ""))
            c.font = _BODY_FONT
            c.fill = qf
            c.border = _BORDER
            c.alignment = _LEFT
            _c(ws, q_row, 7, q.get("marks", ""),  fill=_YELLOW_FILL, bold=True, align=_CENTER)
            _c(ws, q_row, 8, q.get("co_id", ""),  fill=_ORANGE_FILL, bold=True, align=_CENTER)
            bl_raw = str(q.get("bloom_level", ""))
            bl_str = bl_raw if bl_raw.startswith("L") else f"L{bl_raw}" if bl_raw else ""
            _c(ws, q_row, 9, bl_str, fill=_LIME_FILL, bold=True, align=_CENTER)
            ws.row_dimensions[q_row].height = 40
            q_row += 1
    else:
        for i in range(1, 16):
            qf = q_row_fills[i % 2]
            _c(ws, q_row, 1, i, fill=qf, bold=True, align=_CENTER)
            ws.merge_cells(start_row=q_row, start_column=2, end_row=q_row, end_column=6)
            c = ws.cell(row=q_row, column=2, value="")
            c.fill = qf
            c.border = _BORDER
            c.alignment = _LEFT
            _c(ws, q_row, 7, "", fill=_YELLOW_FILL, align=_CENTER)
            _c(ws, q_row, 8, "", fill=_ORANGE_FILL, align=_CENTER)
            _c(ws, q_row, 9, "", fill=_LIME_FILL,   align=_CENTER)
            ws.row_dimensions[q_row].height = 40
            q_row += 1

    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["G"].width = 8
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 6
    ws.column_dimensions["K"].width = 8
    ws.column_dimensions["L"].width = 16
    ws.column_dimensions["M"].width = 12
    ws.column_dimensions["O"].width = 10
    ws.column_dimensions["P"].width = 14
    ws.column_dimensions["Q"].width = 12
    ws.column_dimensions["R"].width = 12
    return ws


# ─────────────────────────────────────────────────────────────────────────────
# Marks sheet (ESE_MKS or CA{n}_Marks)
# ─────────────────────────────────────────────────────────────────────────────
def _build_marks_sheet(wb, sheet_name, course, ca_label, qp_sheet_name,
                       students, total_marks, saved_marks=None, saved_qp=None):
    ws = wb.create_sheet(sheet_name)
    _course_header_block(ws, course)

    n_students = len(students)
    data_start = 10
    data_end   = data_start + n_students - 1
    count_end  = max(data_end, 224)

    # ── Build normalised saved marks lookup FIRST ─────────────────────────
    # key: normalised PRN string → {q_no_str: float} or {"_total": float}
    _saved = {}
    if saved_marks:
        for prn_key, qmarks in saved_marks.items():
            try:
                norm = str(int(float(str(prn_key))))
            except (ValueError, TypeError):
                norm = str(prn_key).strip()
            _saved[norm] = qmarks

    # ── Determine actual question count from QP or marks ──────────────────
    # If we have a QP, use its length. Otherwise infer from marks.
    if saved_qp:
        n_qs = len(saved_qp)
    elif _saved:
        # Find max numeric key across all student marks
        max_q = 0
        for qmarks in _saved.values():
            if isinstance(qmarks, dict):
                for k in qmarks.keys():
                    if k != "_total":
                        try:
                            max_q = max(max_q, int(float(str(k))))
                        except (ValueError, TypeError):
                            pass
        n_qs = max_q if max_q > 0 else 0
    else:
        n_qs = 0   # no QP, no marks → blank template

    has_questions = n_qs > 0

    # ── Row 7: column headers ─────────────────────────────────────────────
    r = 7
    hdr_info  = ["Sr. No.", "Seat No", "Roll No.", "Name of the Student", ca_label]
    hdr_fills = [_NAVY_FILL, _NAVY_FILL, _NAVY_FILL, _NAVY_FILL, _TEAL_FILL]
    for ci, (h, hf) in enumerate(zip(hdr_info, hdr_fills), 1):
        _c(ws, r, ci, h, font=_HEADER_FONT, fill=hf)

    if has_questions:
        # Write hardcoded Q.no headers (not formula-dependent on QP sheet)
        for qi in range(n_qs):
            q_no_val = saved_qp[qi].get("q_no", qi + 1) if saved_qp and qi < len(saved_qp) else qi + 1
            col = 6 + qi
            _c(ws, r, col, q_no_val, font=_HEADER_FONT, fill=_NAVY_FILL)
    else:
        # No questions — leave row 7 question cols blank (they'll be empty template)
        pass

    # ── Row 8: "Marks" label + CO mapping (hardcoded from QP) ────────────
    r = 8
    _c(ws, r, 5, "Marks", fill=_LIGHT_FILL, bold=True, align=_CENTER)
    if has_questions and saved_qp:
        for qi in range(n_qs):
            q = saved_qp[qi] if qi < len(saved_qp) else {}
            col = 6 + qi
            _c(ws, r, col, q.get("co_id", ""), fill=_ORANGE_FILL, bold=True, align=_CENTER)
    elif has_questions:
        for qi in range(n_qs):
            _c(ws, r, 6 + qi, "", fill=_ORANGE_FILL, align=_CENTER)

    # ── Row 9: total marks + per-question max (hardcoded from QP) ────────
    r = 9
    _c(ws, r, 5, total_marks, fill=_GREEN_FILL, bold=True, align=_CENTER)
    if has_questions and saved_qp:
        for qi in range(n_qs):
            q = saved_qp[qi] if qi < len(saved_qp) else {}
            col = 6 + qi
            _c(ws, r, col, q.get("marks", ""), fill=_LIME_FILL, bold=True, align=_CENTER)
    elif has_questions:
        for qi in range(n_qs):
            _c(ws, r, 6 + qi, "", fill=_LIME_FILL, align=_CENTER)

    # ── Rows 10+: students ────────────────────────────────────────────────
    row_fills = [_GREEN_FILL, _LIGHT_FILL]

    for idx, s in enumerate(students, 1):
        r = data_start + idx - 1
        fill = row_fills[idx % 2]
        prn_val = str(s["prn"]) if s["prn"] is not None else ""
        try:
            prn_norm = str(int(float(prn_val))) if prn_val else ""
        except (ValueError, TypeError):
            prn_norm = prn_val

        _c(ws, r, 1, idx,       fill=fill, bold=True, align=_CENTER)
        _c(ws, r, 2, "",        fill=fill, align=_CENTER)
        _c(ws, r, 3, prn_val,   fill=fill, align=_CENTER)
        _c(ws, r, 4, s["name"], fill=fill, align=_LEFT)

        student_marks = _saved.get(prn_norm, {})

        if has_questions:
            # We have per-question columns — write marks per question
            q_vals = []
            for qi in range(n_qs):
                col = 6 + qi
                mark_val = None
                if student_marks and "_total" not in student_marks:
                    # Try q_no as string key, then integer index
                    if saved_qp and qi < len(saved_qp):
                        actual_qno = str(saved_qp[qi].get("q_no", qi + 1))
                        mark_val = student_marks.get(actual_qno)
                    if mark_val is None:
                        mark_val = student_marks.get(str(qi + 1))
                    if mark_val is None:
                        mark_val = student_marks.get(qi + 1)
                    if mark_val is not None:
                        try:
                            mark_val = float(mark_val)
                        except (TypeError, ValueError):
                            mark_val = None
                _c(ws, r, col, mark_val, fill=fill, align=_CENTER)
                if mark_val is not None:
                    q_vals.append(mark_val)

            # Col E: hardcoded sum of actual marks (not a formula)
            # This ensures the value shows even without Excel recalculation
            if q_vals:
                col_e_val = round(sum(q_vals), 4)
            elif student_marks and "_total" in student_marks:
                try:
                    col_e_val = float(student_marks["_total"])
                except (TypeError, ValueError):
                    col_e_val = None
            else:
                # No marks at all — use a SUM formula as fallback
                last_q_col = get_column_letter(6 + n_qs - 1)
                col_e_val = f"=SUM(F{r}:{last_q_col}{r})"
            _c(ws, r, 5, col_e_val, fill=_YELLOW_FILL, bold=True, align=_CENTER)

        else:
            # No question breakdown — write total directly (or blank)
            if student_marks and "_total" in student_marks:
                try:
                    col_e_val = float(student_marks["_total"])
                except (TypeError, ValueError):
                    col_e_val = None
            elif student_marks:
                # Marks exist but in unexpected format — try to sum them
                try:
                    col_e_val = round(sum(
                        float(v) for v in student_marks.values()
                        if v is not None and str(v) not in ("", "None")
                    ), 4)
                except (TypeError, ValueError):
                    col_e_val = None
            else:
                col_e_val = None
            _c(ws, r, 5, col_e_val, fill=_YELLOW_FILL, bold=True, align=_CENTER)

    # ── Summary rows ──────────────────────────────────────────────────────
    s0 = count_end + 3

    cos   = [c["co_id"] for c in course.cos]
    n_cos = len(cos)

    # Use n_qs for summary columns; if 0, show at least 1 blank summary column
    n_sum_cols = max(n_qs, 1)

    _c(ws, s0, 1, "CO No",  bold=True, fill=_NAVY_FILL, font=_HEADER_FONT, align=_CENTER)
    _c(ws, s0, 2, "Level",  bold=True, fill=_NAVY_FILL, font=_HEADER_FONT, align=_CENTER)
    _c(ws, s0, 4, "No of students who attempted",
       bold=True, fill=_TEAL_FILL, align=_LEFT)

    for qi in range(n_sum_cols):
        col = 6 + qi
        col_ltr = get_column_letter(col)
        if has_questions:
            _c(ws, s0, col,
               f'=IF(OR({col_ltr}$8="",{col_ltr}$9=""),"",COUNT({col_ltr}{data_start}:{col_ltr}{count_end}))',
               fill=_LIGHT_FILL, align=_CENTER)
        else:
            _c(ws, s0, col, f'=COUNT(E{data_start}:E{count_end})',
               fill=_LIGHT_FILL, align=_CENTER)

    row_labels    = ["CO No", "Max", "Target", "No. of students scored >= target", "Percentage"]
    summary_fills = [_LIGHT_FILL, _GREEN_FILL, _YELLOW_FILL, _ORANGE_FILL, _SKYBLUE_FILL]

    for i, co_id in enumerate(cos):
        r_co  = s0 + 1 + i
        sfill = summary_fills[i % len(summary_fills)]
        _c(ws, r_co, 1, co_id,
           bold=True, fill=_NAVY_FILL, font=_HEADER_FONT, align=_CENTER)
        _c(ws, r_co, 4,
           row_labels[i] if i < len(row_labels) else "",
           bold=True, fill=sfill, align=_LEFT)

        for qi in range(n_sum_cols):
            col     = 6 + qi
            col_ltr = get_column_letter(col)
            if has_questions:
                guard = f'IF(OR({col_ltr}$8="",{col_ltr}$9=""),"",{{}})'
            else:
                guard = "{}"  # no guard needed when using col E directly

            if i == 0:  # CO No row
                val = f'=IF(A{r_co}="","",{col_ltr}8)' if has_questions else f'=CO_List!A{14+0}'
                _c(ws, r_co, col, val, fill=sfill, align=_CENTER)
            elif i == 1:  # Max marks
                val = f'={col_ltr}9' if has_questions else f'=E9'
                _c(ws, r_co, col, val, fill=sfill, align=_CENTER)
            elif i == 2:  # Target
                if has_questions:
                    _c(ws, r_co, col,
                       f'=IF(OR({col_ltr}$8="",{col_ltr}$9=""),"",{col_ltr}{r_co-1}*VLOOKUP({col_ltr}{r_co-2},CO_List!$A$14:$J$19,10,0)/100)',
                       fill=sfill, align=_CENTER)
                else:
                    _c(ws, r_co, col,
                       f'=E9*VLOOKUP(E{r_co-2},CO_List!$A$14:$J$19,10,0)/100',
                       fill=sfill, align=_CENTER)
            elif i == 3:  # No. students >= target
                if has_questions:
                    _c(ws, r_co, col,
                       f'=IF(OR({col_ltr}$8="",{col_ltr}$9=""),"",COUNTIFS({col_ltr}{data_start}:{col_ltr}{count_end},">="&{col_ltr}{r_co-1}))',
                       fill=sfill, align=_CENTER)
                else:
                    _c(ws, r_co, col,
                       f'=COUNTIFS(E{data_start}:E{count_end},">="&E{r_co-1})',
                       fill=sfill, align=_CENTER)
            elif i == 4:  # Percentage
                if has_questions:
                    _c(ws, r_co, col,
                       f'=IF(OR({col_ltr}$8="",{col_ltr}$9=""),"",IFERROR({col_ltr}{r_co-1}/{col_ltr}{s0}*100,0))',
                       fill=sfill, align=_CENTER, number_format="0.00")
                else:
                    _c(ws, r_co, col,
                       f'=IFERROR(E{r_co-1}/E{s0}*100,0)',
                       fill=sfill, align=_CENTER, number_format="0.00")

    # CO attainment level in col B — compute in Python and hardcode
    # This ensures Final_CO_Attn can read them without formula recalculation
    pct_row      = s0 + 5
    co_row_start = s0 + 1

    # Collect per-CO percentage values (computed above in summary rows)
    # We need to derive the level from the actual student marks data
    co_levels = {}
    for i, co_id in enumerate(cos):
        r_co = s0 + 1 + i
        # Compute percentage from marks in Python
        if has_questions and saved_qp and any(q.get("co_id") for q in saved_qp):
            # Find which question columns map to this CO
            co_q_cols = []
            for qi, q in enumerate(saved_qp):
                if q.get("co_id") == co_id:
                    co_q_cols.append(qi)
            if not co_q_cols:
                # No QP mapping, try all questions
                co_q_cols = list(range(n_qs))
            # Max marks for this CO
            max_co = sum(
                (saved_qp[qi].get("marks") or 0) for qi in co_q_cols
                if qi < len(saved_qp)
            )
            if max_co <= 0:
                co_levels[co_id] = None
                continue
            target = max_co * 0.60
            n_total = len([s for s in students if s.get("prn")])
            if n_total == 0:
                co_levels[co_id] = None
                continue
            n_scored = 0
            for s in students:
                prn_val = str(s["prn"]) if s["prn"] is not None else ""
                try:
                    prn_norm = str(int(float(prn_val))) if prn_val else ""
                except (ValueError, TypeError):
                    prn_norm = prn_val
                sm = _saved.get(prn_norm, {})
                student_co_total = 0
                for qi in co_q_cols:
                    if saved_qp and qi < len(saved_qp):
                        actual_qno = str(saved_qp[qi].get("q_no", qi + 1))
                        v = sm.get(actual_qno) or sm.get(str(qi + 1)) or sm.get(qi + 1)
                    else:
                        v = sm.get(str(qi + 1)) or sm.get(qi + 1)
                    if v is not None:
                        try:
                            student_co_total += float(v)
                        except (TypeError, ValueError):
                            pass
                if student_co_total >= target:
                    n_scored += 1
            pct = (n_scored / n_total) * 100 if n_total > 0 else 0
            level = 3 if pct >= 70 else (2 if pct > 40 else 1)
            co_levels[co_id] = level
        elif not has_questions or not saved_qp or not any(q.get("co_id") for q in (saved_qp or [])):
            # No question breakdown — use overall total marks
            target = (total_marks or 0) * 0.60
            n_total = len([s for s in students if s.get("prn")])
            if n_total == 0:
                co_levels[co_id] = None
                continue
            n_scored = 0
            for s in students:
                prn_val = str(s["prn"]) if s["prn"] is not None else ""
                try:
                    prn_norm = str(int(float(prn_val))) if prn_val else ""
                except (ValueError, TypeError):
                    prn_norm = prn_val
                sm = _saved.get(prn_norm, {})
                val = None
                if "_total" in sm:
                    try:
                        val = float(sm["_total"])
                    except (TypeError, ValueError):
                        pass
                elif sm:
                    try:
                        val = round(sum(float(v) for v in sm.values()
                                        if v is not None and str(v) not in ("", "None")), 4)
                    except (TypeError, ValueError):
                        pass
                if val is not None and target > 0 and val >= target:
                    n_scored += 1
            pct = (n_scored / n_total) * 100 if n_total > 0 else 0
            level = 3 if pct >= 70 else (2 if pct > 40 else 1)
            co_levels[co_id] = level
        else:
            co_levels[co_id] = None

    for i, co_id in enumerate(cos):
        r_co  = s0 + 1 + i
        sfill = summary_fills[i % len(summary_fills)]
        level_val = co_levels.get(co_id)
        _c(ws, r_co, 2, level_val, fill=_YELLOW_FILL, bold=True, align=_CENTER)

    # Level row
    r_level = s0 + 1 + n_cos
    _c(ws, r_level, 4, "Level", bold=True, fill=_TEAL_FILL, align=_LEFT)
    for qi in range(n_sum_cols):
        col     = 6 + qi
        col_ltr = get_column_letter(col) if has_questions else "E"
        _c(ws, r_level, col,
           f'=IF(OR({col_ltr}$8="",{col_ltr}$9=""),'
           f'"",IF({col_ltr}{pct_row}>=70,3,IF({col_ltr}{pct_row}>40,2,IF({col_ltr}{pct_row}<=40,1,0))))',
           fill=_SKYBLUE_FILL, bold=True, align=_CENTER)

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 12

    return {"s0": s0, "co_rows": [s0 + 1 + i for i in range(n_cos)], "co_levels": co_levels}


# ─────────────────────────────────────────────────────────────────────────────
# Final_CO_Attn
# ─────────────────────────────────────────────────────────────────────────────
def _build_final_co_attn(wb, course, ca_names, marks_meta, ese_meta):
    ws = wb.create_sheet("Final_CO_Attn")
    _course_header_block(ws, course)

    cos  = [c["co_id"] for c in course.cos]
    n_ca = len(ca_names)

    r = 7
    _c(ws, r, 1, "CO No / Weightage",
       fill=_NAVY_FILL, font=_HEADER_FONT, bold=True, align=_CENTER)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=n_ca + 1)
    c = ws.cell(row=r, column=2, value="CO Attainment using CIE")
    c.font = _HEADER_FONT; c.fill = _TEAL_FILL; c.border = _BORDER; c.alignment = _CENTER
    ws.merge_cells(start_row=r, start_column=n_ca + 2, end_row=r, end_column=n_ca + 4)
    c = ws.cell(row=r, column=n_ca + 2, value="Final CO attainment")
    c.font = _HEADER_FONT; c.fill = _ORANGE_FILL; c.border = _BORDER; c.alignment = _CENTER
    _c(ws, r, n_ca + 5, "Overall Att",
       fill=_LIME_FILL, font=Font(name="Calibri", bold=True, size=10), bold=True)

    r = 8
    _c(ws, r, 1, "", fill=_LIGHT_FILL, align=_CENTER)
    for i, ca in enumerate(ca_names):
        _c(ws, r, i + 2, ca, bold=True, fill=_SKYBLUE_FILL, align=_CENTER)
    g = n_ca + 2
    _c(ws, r, g,     "Internal", bold=True, fill=_GREEN_FILL,  align=_CENTER)
    _c(ws, r, g + 1, "External", bold=True, fill=_ORANGE_FILL, align=_CENTER)
    _c(ws, r, g + 2, "Final",    bold=True, fill=_TEAL_FILL,   align=_CENTER)

    r = 9
    _c(ws, r, 1, "", fill=_LIGHT_FILL, align=_CENTER)
    for i in range(n_ca):
        _c(ws, r, i + 2, "", fill=_LIGHT_FILL, align=_CENTER)
    _c(ws, r, g,     40,  fill=_GREEN_FILL,  bold=True, align=_CENTER)
    _c(ws, r, g + 1, 60,  fill=_ORANGE_FILL, bold=True, align=_CENTER)
    _c(ws, r, g + 2, 100, fill=_TEAL_FILL,   bold=True, align=_CENTER)
    wt_row = r

    r = 10
    co_fills = [_GREEN_FILL, _LIGHT_FILL, _YELLOW_FILL, _ORANGE_FILL, _PINK_FILL]
    internal_levels = []  # list of per-CO internal (CIE) level values
    final_co_values = {}  # co_id -> final attainment value (for PO_Attainment)
    for ci, co_id in enumerate(cos):
        cfill = co_fills[ci % len(co_fills)]
        # Hardcode CO ID (avoid formula referencing CO_List which may be uncalculated)
        _c(ws, r, 1, co_id, fill=_NAVY_FILL, font=_HEADER_FONT, bold=True, align=_CENTER)

        # CA levels: hardcoded from co_levels computed in Python
        ca_level_vals = []
        for cai, ca in enumerate(ca_names):
            level_val = marks_meta[cai].get("co_levels", {}).get(co_id)
            _c(ws, r, cai + 2, level_val, fill=cfill, bold=True, align=_CENTER)
            if level_val is not None:
                ca_level_vals.append(level_val)

        # Internal (CIE) = average of CA levels
        if ca_level_vals:
            internal_val = round(sum(ca_level_vals) / len(ca_level_vals), 2)
        else:
            internal_val = None
        _c(ws, r, g, internal_val, fill=_GREEN_FILL, bold=True, align=_CENTER,
           number_format="0.00")
        internal_levels.append(internal_val)

        # External (ESE) level
        ese_level_val = ese_meta.get("co_levels", {}).get(co_id)
        _c(ws, r, g + 1, ese_level_val, fill=_ORANGE_FILL, bold=True, align=_CENTER,
           number_format="0.00")

        # Final = Internal*0.4 + External*0.6
        int_v = internal_val or 0
        ext_v = ese_level_val or 0
        if internal_val is not None or ese_level_val is not None:
            final_val = round(int_v * 0.40 + ext_v * 0.60, 2)
        else:
            final_val = None
        _c(ws, r, g + 2, final_val, fill=_TEAL_FILL, bold=True, align=_CENTER,
           number_format="0.00")
        final_co_values[co_id] = final_val

        r += 1

    final_col = get_column_letter(g + 2)
    # Overall CO Attainment — hardcoded average of final CO attainment values
    final_vals_list = [v for v in final_co_values.values() if v is not None]
    overall_val = round(sum(final_vals_list) / len(final_vals_list), 2) if final_vals_list else None
    _c(ws, r, 1, "Overall CO Attainment",
       bold=True, fill=_SKYBLUE_FILL, align=_LEFT)
    _c(ws, r, g + 4, overall_val,
       fill=_LIME_FILL, bold=True, align=_CENTER, number_format="0.00")
    r += 2

    _c(ws, r, 1, "Final CO attainment", bold=True, fill=_LIGHT_FILL, align=_LEFT)
    r += 1
    _c(ws, r, 1, "External"); _c(ws, r, 2, 0.6, fill=_YELLOW_FILL, bold=True)
    _c(ws, r, 3, "OR"); _c(ws, r, 4, 1.0, fill=_GREEN_FILL, bold=True)
    _c(ws, r, 5, "OR"); _c(ws, r, 6, "Nil", fill=_ORANGE_FILL)
    r += 1
    _c(ws, r, 1, "Internal"); _c(ws, r, 2, 0.4, fill=_YELLOW_FILL, bold=True)
    _c(ws, r, 4, "NIL", fill=_LIGHT_FILL); _c(ws, r, 6, 1.0, fill=_GREEN_FILL, bold=True)
    r += 1
    _c(ws, r, 2, "Both", fill=_GREEN_FILL)
    _c(ws, r, 4, "Only ESE", fill=_ORANGE_FILL)
    _c(ws, r, 6, "Only CIE", fill=_SKYBLUE_FILL)

    ws.column_dimensions["A"].width = 20
    for i in range(n_ca + 6):
        ws.column_dimensions[get_column_letter(i + 2)].width = 12

    return final_co_values


# ─────────────────────────────────────────────────────────────────────────────
# PO_Attainment
# ─────────────────────────────────────────────────────────────────────────────
def _build_po_attainment(wb, course, n_ca, final_co_values=None):
    ws = wb.create_sheet("PO_Attainment")
    _course_header_block(ws, course)

    co_po  = course.co_po_matrix
    cos    = [c["co_id"] for c in course.cos]
    pos    = course.pos
    po_ids = [p["po_id"] for p in pos] if pos else [f"PO{i}" for i in range(1, 13)]
    final_co_values = final_co_values or {}

    r = 7
    _c(ws, r, 1, "CO",         bold=True, fill=_NAVY_FILL, font=_HEADER_FONT)
    _c(ws, r, 2, "Attainment", bold=True, fill=_NAVY_FILL, font=_HEADER_FONT)
    for ci, po in enumerate(po_ids, 3):
        _c(ws, r, ci, po, bold=True, fill=_NAVY_FILL, font=_HEADER_FONT)

    r = 8
    data_start = r

    co_fills = [_GREEN_FILL, _LIGHT_FILL, _YELLOW_FILL, _ORANGE_FILL, _PINK_FILL]
    for ci_idx, co_id in enumerate(cos):
        cfill = co_fills[ci_idx % len(co_fills)]
        _c(ws, r, 1, co_id, bold=True, fill=cfill, align=_CENTER)
        # Hardcode final CO attainment value instead of cross-sheet formula
        final_val = final_co_values.get(co_id)
        _c(ws, r, 2, final_val, fill=_YELLOW_FILL, bold=True, align=_CENTER, number_format="0.00")
        mapping = co_po.get(co_id, {})
        for pi, po in enumerate(po_ids, 3):
            val = mapping.get(po, None)
            cell_fill = _GREEN_FILL if val else PatternFill()
            _c(ws, r, pi, val if val else "", fill=cell_fill,
               align=_CENTER, bold=bool(val))
        r += 1

    # Articulation Average — hardcoded average of non-empty CO-PO mapping values per PO
    _c(ws, r, 1, "Articulation Average", bold=True, fill=_TEAL_FILL, align=_LEFT)
    _c(ws, r, 2, "", fill=_TEAL_FILL)
    for pi, po in enumerate(po_ids, 3):
        vals = []
        for co_id in cos:
            v = co_po.get(co_id, {}).get(po)
            if v:
                try:
                    vals.append(float(v))
                except (TypeError, ValueError):
                    pass
        avg = round(sum(vals) / len(vals), 2) if vals else "-"
        _c(ws, r, pi, avg, fill=_ORANGE_FILL, bold=True, align=_CENTER, number_format="0.00")
    r += 1

    # CO-PO_PSO Attainment — hardcoded: sum(mapping * co_attainment) / (3 * mapped_count)
    _c(ws, r, 1, "CO-PO_PSO Attainment", bold=True, fill=_LIME_FILL, align=_LEFT)
    _c(ws, r, 2, "", fill=_LIME_FILL)
    for pi, po in enumerate(po_ids, 3):
        numerator = 0.0
        mapped_count = 0
        for co_id in cos:
            mapping_val = co_po.get(co_id, {}).get(po)
            attn_val = final_co_values.get(co_id)
            if mapping_val and attn_val is not None:
                try:
                    numerator += float(mapping_val) * float(attn_val)
                    mapped_count += 1
                except (TypeError, ValueError):
                    pass
        result = round(numerator / (3 * mapped_count), 2) if mapped_count > 0 else "-"
        _c(ws, r, pi, result, fill=_SKYBLUE_FILL, bold=True, align=_CENTER, number_format="0.00")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    for pi in range(3, 3 + len(po_ids)):
        ws.column_dimensions[get_column_letter(pi)].width = 8


# ─────────────────────────────────────────────────────────────────────────────
# Main service
# ─────────────────────────────────────────────────────────────────────────────
class COPOTemplateService:
    def __init__(self, db: AsyncSession):
        self.db = db

    @staticmethod
    def get_filepath(course_id: int) -> str:
        storage = get_storage()
        p = storage.get_path(_CATEGORY, f"co_po_template_{course_id}.xlsx")
        return str(p) if p else str(
            get_storage()._dir(_CATEGORY) / f"co_po_template_{course_id}.xlsx"
        )

    async def _get_students(self, course_id: int):
        try:
            result = await self.db.execute(
                text("SELECT prn, name, section FROM students WHERE course_id=:cid ORDER BY section, name"),
                {"cid": course_id}
            )
            rows = result.fetchall()
            return [{"prn": r[0], "name": r[1], "section": r[2]} for r in rows]
        except Exception as e:
            logger.warning(f"Could not fetch students: {e}")
            return []

    async def _get_questions(self, course_id: int):
        from sqlalchemy import select
        from backend.database.models import Question
        result = await self.db.execute(
            select(Question).where(Question.course_id == course_id)
        )
        return result.scalars().all()

    async def _get_saved_sheets(self, course_id: int) -> dict:
        """Load all saved CASheet records. Returns {ca_label: {qp:[...], marks:{...}}}."""
        from sqlalchemy import select
        from backend.database.models import CASheet
        try:
            result = await self.db.execute(
                select(CASheet).where(CASheet.course_id == course_id)
            )
            return {s.ca_label: {"qp": s.qp, "marks": s.marks}
                    for s in result.scalars().all()}
        except Exception as e:
            logger.warning(f"Could not load saved CA sheets: {e}")
            return {}

    # All names that mean "End Semester Exam" — mapped to canonical key "ese"
    _ESE_ALIASES = {
        "ese", "end semester", "end-semester", "end sem", "end-sem",
        "final exam", "final examination", "semester exam", "semester examination",
        "end semester exam", "end semester examination", "endsem",
    }

    @classmethod
    def _norm_label(cls, label: str) -> str:
        """Lowercase + map all ESE aliases → 'ese'. Otherwise return as-is lowercase."""
        s = label.lower().strip()
        if s in cls._ESE_ALIASES:
            return "ese"
        for suffix in (" exam", " examination"):
            if s.endswith(suffix):
                base = s[: -len(suffix)].strip()
                if base in cls._ESE_ALIASES:
                    return "ese"
        return s

    @classmethod
    def _norm_base(cls, label: str) -> str:
        """Strip trailing number: 'Unit Test 2' → 'unit test', 'Quiz 1' → 'quiz'."""
        import re as _re2
        s = cls._norm_label(label)
        return _re2.sub(r'\s*\d+\s*$', '', s).strip()

    async def _get_attainment_marks(self, course_id: int) -> dict:
        """
        Load COAttainment records and return per-component marks as:
            { component_name: { prn: {"_total": float} OR {q_no: float, ...} } }

        Handles all storage formats produced by parse_marks_xlsx:
          co_wise        — {CO1: {Quiz: 8.0, UT: 14.0}, CO2: {Quiz: 5.0, ...}}
          component_wise — {Quiz: {Total: 8.5, 1: 2.5, 2: 3.0, ...}}
          exam_wise_flat — {Quiz: 8.5, UT: 14.0}

        For co_wise: a student's total for component X = sum of X across all COs.
        This is correct because CO1.Quiz + CO2.Quiz = all marks they got in Quiz.
        """
        from sqlalchemy import select
        from backend.database.models import COAttainment
        try:
            result = await self.db.execute(
                select(COAttainment).where(COAttainment.course_id == course_id)
            )
            records = result.scalars().all()
        except Exception as e:
            logger.warning(f"Could not load COAttainment records: {e}")
            return {}

        if not records:
            return {}

        # Detect format from first record
        sample = records[0].marks or {}
        first_val = next(iter(sample.values()), None) if sample else None

        if isinstance(first_val, dict):
            first_inner = next(iter(first_val.values()), None)
            if isinstance(first_inner, dict):
                fmt = "co_wise"        # {CO1: {Quiz: 8, UT: 14}}
            else:
                fmt = "component_wise" # {Quiz: {Total: 8.5, 1: 2.5}}
        else:
            fmt = "exam_wise_flat"     # {Quiz: 8.5, UT: 14}

        comp_marks: dict = {}

        for rec in records:
            # Normalise PRN key the same way as students table
            prn = str(rec.student_id).strip()
            try:
                prn = str(int(float(prn)))
            except (ValueError, TypeError):
                pass

            marks = rec.marks or {}

            if fmt == "co_wise":
                # Sum each component's value across all COs
                comp_totals: dict = {}
                for co_id, co_dict in marks.items():
                    if not isinstance(co_dict, dict):
                        continue
                    for comp, val in co_dict.items():
                        try:
                            comp_totals[comp] = comp_totals.get(comp, 0.0) + float(val)
                        except (TypeError, ValueError):
                            pass
                for comp, total in comp_totals.items():
                    comp_marks.setdefault(comp, {})[prn] = {"_total": round(total, 4)}

            elif fmt == "component_wise":
                # {comp: {Total: 8.5, 1: 2.5, 2: 3.0, ...}} or {comp: {Total: 8.5, Q1: 2.5}}
                for comp, val in marks.items():
                    if isinstance(val, dict):
                        # Separate "Total" from question-level marks
                        q_marks = {}
                        total_val = None
                        for k, v in val.items():
                            if str(k).lower() == "total":
                                try:
                                    total_val = float(v)
                                except (TypeError, ValueError):
                                    pass
                            else:
                                # Convert Q1/Q2 keys to "1"/"2"
                                import re as _re
                                m = _re.match(r'Q?(\d+)', str(k), _re.IGNORECASE)
                                q_key = m.group(1) if m else str(k)
                                try:
                                    q_marks[q_key] = float(v)
                                except (TypeError, ValueError):
                                    pass
                        if q_marks:
                            comp_marks.setdefault(comp, {})[prn] = q_marks
                        elif total_val is not None:
                            comp_marks.setdefault(comp, {})[prn] = {"_total": total_val}
                    else:
                        try:
                            comp_marks.setdefault(comp, {})[prn] = {"_total": float(val)}
                        except (TypeError, ValueError):
                            pass

            else:  # exam_wise_flat
                for comp, val in marks.items():
                    try:
                        comp_marks.setdefault(comp, {})[prn] = {"_total": float(val)}
                    except (TypeError, ValueError):
                        pass

        return comp_marks

    async def generate(self, course_id: int, qp_source: str = "blank") -> dict:
        course_svc = CourseService(self.db)
        course     = await course_svc.get_course(course_id)
        students   = await self._get_students(course_id)
        eval_cfg   = course.evaluation_config
        components = eval_cfg.get("components", {})

        # Load saved CA sheets (QP + per-question marks entered in the frontend)
        saved_sheets = await self._get_saved_sheets(course_id)
        # Load COAttainment marks (from mark upload page)
        attainment_marks = await self._get_attainment_marks(course_id)

        logger.info(f"[generate] course_id={course_id} students={len(students)} "
                    f"saved_sheets={list(saved_sheets.keys())} "
                    f"attainment_comps={list(attainment_marks.keys())}")

        # Build CA name list — merge all three sources with smart deduplication:
        # - Numbered variants like "Unit Test 1" and "Unit Test 2" are DIFFERENT → both kept
        # - Only deduplicate when norm labels are identical (e.g. "quiz" == "quiz")
        # - ESE variants all map to "ese" → excluded from CA list regardless of name
        ESE_KEYWORDS = {"end semester", "ese", "end-semester", "final exam", "end sem",
                        "endsem", "end sem", "semester exam", "final examination"}
        def _is_ese(name):
            return self._norm_label(name) == "ese"

        comp_ca_names  = sorted([k for k in components.keys() if not _is_ese(k)])
        saved_ca_names = [k for k in saved_sheets.keys()    if not _is_ese(k)]
        attn_ca_names  = [k for k in attainment_marks.keys() if not _is_ese(k)]

        # Deduplicate by exact normalised label (NOT by base — keep numbered variants)
        _seen_norm: set = set()
        ca_names: list = []

        def _add_if_new(label):
            nk = self._norm_label(label)
            if nk not in _seen_norm:
                _seen_norm.add(nk)
                ca_names.append(label)

        # Priority order: CASheet > COAttainment > eval_config
        for label in saved_ca_names:
            _add_if_new(label)
        for label in attn_ca_names:
            _add_if_new(label)
        for label in comp_ca_names:
            _add_if_new(label)

        if not ca_names:
            ca_names = [f"CA{i}" for i in range(1, 4)]

        # Final cleanup: remove generic eval_config names that are base-subsumed
        # e.g. if "Quiz 1" already in list, drop "Quiz" (its base norm "quiz" is covered)
        _covered_bases = {self._norm_base(n) for n in ca_names}
        ca_names = [
            n for n in ca_names
            if self._norm_label(n) not in _covered_bases   # not a pure-base duplicate
            or not any(
                self._norm_base(other) == self._norm_label(n) and other != n
                for other in ca_names
            )
        ]

        ca_names = ca_names[:12]  # safety cap

        wb = Workbook()
        del wb[wb.sheetnames[0]]

        # 1. Course_Info
        _build_course_info(wb, course)

        # 2. Roll_List
        _build_roll_list(wb, course, students)

        # 3. CO_List
        _build_co_list(wb, course)

        # Build normalised lookups for attainment_marks
        # exact_norm: normalised label → actual key  (for exact match after ESE canonicalisation)
        # base_norm:  base label (no trailing number) → list of actual keys  (for fuzzy grouping)
        _exact_norm: dict = {}
        _base_norm: dict = {}
        for ak in attainment_marks.keys():
            nk = self._norm_label(ak)
            _exact_norm[nk] = ak
            bk = self._norm_base(ak)
            _base_norm.setdefault(bk, []).append(ak)

        def _find_attn_key(label: str) -> dict:
            """Return attainment marks for label using exact → ESE-alias → base-fuzzy match."""
            # 1. Exact match
            if label in attainment_marks:
                return attainment_marks[label]
            # 2. Normalised exact (handles ESE aliases: "End Semester" → "ese" == "ESE" → "ese")
            nk = self._norm_label(label)
            if nk in _exact_norm:
                return attainment_marks[_exact_norm[nk]]
            # 3. Base fuzzy: "Quiz" matches "Quiz 1", "Unit Test" matches "Unit Test 1"+"Unit Test 2"
            bk = self._norm_base(label)
            matches = _base_norm.get(bk, [])
            if not matches:
                # also try exact norm as base key
                matches = _base_norm.get(nk, [])
            if matches:
                merged_attn: dict = {}
                for ak in matches:
                    for prn, mks in attainment_marks[ak].items():
                        if prn not in merged_attn:
                            merged_attn[prn] = mks
                        elif "_total" in merged_attn[prn] and "_total" not in mks:
                            merged_attn[prn] = mks
                return merged_attn
            return {}

        # ── Helper: merge CASheet marks + COAttainment marks for a component ──
        def _merged_marks(ca_label):
            """
            Returns (saved_qp, merged_marks_dict) for a component.
            CASheet data wins over COAttainment data (more granular).
            Uses fuzzy matching for COAttainment keys.
            """
            ca_saved     = saved_sheets.get(ca_label, {})
            ca_saved_qp  = ca_saved.get("qp") or []
            ca_sheet_mks = ca_saved.get("marks") or {}
            ca_attn_mks  = _find_attn_key(ca_label)

            # Merge: attainment as base, CASheet on top
            merged = {**ca_attn_mks}
            for prn, mks in ca_sheet_mks.items():
                try:
                    norm = str(int(float(str(prn))))
                except (ValueError, TypeError):
                    norm = str(prn).strip()
                merged[norm] = mks

            return ca_saved_qp or None, merged if merged else None

        # 4. ESE sheets
        ese_saved_qp, ese_saved_marks = _merged_marks("ESE")

        # Fall back to question bank if no QP saved and qp_source says so
        ese_questions = ese_saved_qp
        if not ese_questions and qp_source == "question_bank":
            qs = await self._get_questions(course_id)
            ese_questions = [
                {"q_no": i + 1, "question_text": q.question_text,
                 "marks": q.marks, "co_id": q.co_id, "bloom_level": q.bloom_level}
                for i, q in enumerate(qs[:20])
            ]

        _build_qp_sheet(wb, "ESE_QP", course, "ESE", questions=ese_questions)
        ese_total = eval_cfg.get("end_sem_total", 60)
        ese_meta  = _build_marks_sheet(
            wb, "ESE_MKS", course, "ESE", "ESE_QP", students, ese_total,
            saved_marks=ese_saved_marks,
            saved_qp=ese_saved_qp if ese_saved_qp else ese_questions,
        )

        # 5. CA sheets
        marks_meta_list = []
        for ca in ca_names:
            qp_name  = f"{ca}_QP"
            mks_name = f"{ca}_Marks"

            ca_saved_qp, ca_saved_marks = _merged_marks(ca)

            ca_questions = ca_saved_qp
            if not ca_questions and qp_source == "question_bank":
                qs = await self._get_questions(course_id)
                ca_questions = [
                    {"q_no": i + 1, "question_text": q.question_text,
                     "marks": q.marks, "co_id": q.co_id, "bloom_level": q.bloom_level}
                    for i, q in enumerate(qs[:15])
                ]

            _build_qp_sheet(wb, qp_name, course, ca, questions=ca_questions)

            ca_total = components.get(ca, 10)
            if not isinstance(ca_total, (int, float)):
                ca_total = 10

            meta = _build_marks_sheet(
                wb, mks_name, course, ca, qp_name, students, ca_total,
                saved_marks=ca_saved_marks,
                saved_qp=ca_saved_qp if ca_saved_qp else ca_questions,
            )
            marks_meta_list.append(meta)

        # 6. Final_CO_Attn
        final_co_values = _build_final_co_attn(wb, course, ca_names, marks_meta_list, ese_meta)

        # 7. PO_Attainment
        _build_po_attainment(wb, course, len(ca_names), final_co_values=final_co_values)

        # Save
        import tempfile
        _storage  = get_storage()
        _filename = f"co_po_template_{course_id}.xlsx"
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp) / _filename
            wb.save(str(tmp_path))
            _storage.save_from_path(_CATEGORY, _filename, tmp_path)

        filepath = str(_storage.get_path(_CATEGORY, _filename))
        logger.info(f"CO-PO template saved -> {filepath}")

        return {
            "course_id":    course_id,
            "course_name":  course.course_name,
            "filename":     _filename,
            "download_url": f"/co-po-template/download/{course_id}",
            "sheets":       wb.sheetnames,
            "students":     len(students),
            "ca_count":     len(ca_names),
        }
